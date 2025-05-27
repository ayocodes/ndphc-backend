from datetime import datetime, date
from typing import Optional
from fastapi import APIRouter, Depends, HTTPException, Response
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from io import BytesIO
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy import asc

from app.api.deps import get_db, get_current_user
from app.models.user import User
from app.models.daily_report import DailyReport, TurbineDailyStats, TurbineHourlyGeneration
from app.schemas.user import UserRole

router = APIRouter()

def create_excel_report(
    daily_reports: list[DailyReport],
    turbine_stats: list[TurbineDailyStats],
    hourly_generations: list[TurbineHourlyGeneration]
) -> BytesIO:
    """Create an Excel file with multiple sheets containing the report data."""
    output = BytesIO()
    wb = Workbook()
    
    # Sort data by date (ascending)
    daily_reports.sort(key=lambda x: x.date)
    turbine_stats.sort(key=lambda x: (x.daily_report.date, x.turbine.name))
    hourly_generations.sort(key=lambda x: (x.daily_report.date, x.turbine.name, x.hour))
    
    # Daily Reports Sheet
    ws_daily = wb.active
    ws_daily.title = "Daily Reports"
    headers = [
        "Date", "Power Plant", "Gas Loss (MWh)", "NCC Loss (MWh)", 
        "Internal Loss (MWh)", "Gas Consumed (MMSCH)", "Declaration Total (MW)",
        "Availability Capacity (MW)", "Availability Factor (%)", "Plant Heat Rate",
        "Thermal Efficiency (%)", "Energy Generated (MWh)", "Total Energy Exported (MWh)",
        "Energy Consumed (MWh)", "Availability Forecast (MWh)", "Dependability Index (%)",
        "Avg Energy Sent Out (MW)", "Gas Utilization (MWh/MSCM)", "Load Factor (%)"
    ]
    
    # Style headers
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    for col, header in enumerate(headers, 1):
        cell = ws_daily.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
    
    # Add data
    for row, report in enumerate(daily_reports, 2):
        ws_daily.cell(row=row, column=1, value=report.date)
        ws_daily.cell(row=row, column=2, value=report.power_plant.name)
        ws_daily.cell(row=row, column=3, value=float(report.gas_loss))
        ws_daily.cell(row=row, column=4, value=float(report.ncc_loss))
        ws_daily.cell(row=row, column=5, value=float(report.internal_loss))
        ws_daily.cell(row=row, column=6, value=float(report.gas_consumed))
        ws_daily.cell(row=row, column=7, value=float(report.declaration_total) if report.declaration_total else None)
        ws_daily.cell(row=row, column=8, value=float(report.availability_capacity) if report.availability_capacity else None)
        ws_daily.cell(row=row, column=9, value=float(report.availability_factor) if report.availability_factor else None)
        ws_daily.cell(row=row, column=10, value=float(report.plant_heat_rate) if report.plant_heat_rate else None)
        ws_daily.cell(row=row, column=11, value=float(report.thermal_efficiency) if report.thermal_efficiency else None)
        ws_daily.cell(row=row, column=12, value=float(report.energy_generated) if report.energy_generated else None)
        ws_daily.cell(row=row, column=13, value=float(report.total_energy_exported) if report.total_energy_exported else None)
        ws_daily.cell(row=row, column=14, value=float(report.energy_consumed) if report.energy_consumed else None)
        ws_daily.cell(row=row, column=15, value=float(report.availability_forecast) if report.availability_forecast else None)
        ws_daily.cell(row=row, column=16, value=float(report.dependability_index) if report.dependability_index else None)
        ws_daily.cell(row=row, column=17, value=float(report.avg_energy_sent_out) if report.avg_energy_sent_out else None)
        ws_daily.cell(row=row, column=18, value=float(report.gas_utilization) if report.gas_utilization else None)
        ws_daily.cell(row=row, column=19, value=float(report.load_factor) if report.load_factor else None)
    
    # Turbine Stats Sheet
    ws_turbine = wb.create_sheet("Turbine Stats")
    headers = [
        "Date", "Turbine", "Energy Generated (MWh)", "Energy Exported (MWh)",
        "Operating Hours", "Startup Count", "Shutdown Count", "Trips"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws_turbine.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
    
    for row, stat in enumerate(turbine_stats, 2):
        ws_turbine.cell(row=row, column=1, value=stat.daily_report.date)
        ws_turbine.cell(row=row, column=2, value=stat.turbine.name)
        ws_turbine.cell(row=row, column=3, value=float(stat.energy_generated))
        ws_turbine.cell(row=row, column=4, value=float(stat.energy_exported))
        ws_turbine.cell(row=row, column=5, value=float(stat.operating_hours))
        ws_turbine.cell(row=row, column=6, value=stat.startup_count)
        ws_turbine.cell(row=row, column=7, value=stat.shutdown_count)
        ws_turbine.cell(row=row, column=8, value=stat.trips)
    
    # Hourly Generation Sheet
    ws_hourly = wb.create_sheet("Hourly Generation")
    headers = ["Date", "Turbine", "Hour", "Energy Generated (MW)"]
    
    for col, header in enumerate(headers, 1):
        cell = ws_hourly.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
    
    for row, gen in enumerate(hourly_generations, 2):
        ws_hourly.cell(row=row, column=1, value=gen.daily_report.date)
        ws_hourly.cell(row=row, column=2, value=gen.turbine.name)
        ws_hourly.cell(row=row, column=3, value=gen.hour)
        ws_hourly.cell(row=row, column=4, value=float(gen.energy_generated))
    
    # Adjust column widths
    for ws in [ws_daily, ws_turbine, ws_hourly]:
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width
    
    wb.save(output)
    output.seek(0)
    return output

@router.get("/download")
async def download_data(
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    power_plant_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Download plant data in Excel format. Only accessible by admin users.
    """
    if current_user.role != UserRole.ADMIN:
        raise HTTPException(
            status_code=403,
            detail="Only admin users can download data"
        )
    
    # Build query with default sorting by date ascending
    query = db.query(DailyReport).order_by(asc(DailyReport.date))
    
    if start_date:
        query = query.filter(DailyReport.date >= start_date)
    if end_date:
        query = query.filter(DailyReport.date <= end_date)
    if power_plant_id:
        query = query.filter(DailyReport.power_plant_id == power_plant_id)
    
    # Get all related data
    daily_reports = query.all()
    report_ids = [report.id for report in daily_reports]
    
    turbine_stats = db.query(TurbineDailyStats).filter(
        TurbineDailyStats.daily_report_id.in_(report_ids)
    ).order_by(asc(TurbineDailyStats.daily_report_id)).all()
    
    hourly_generations = db.query(TurbineHourlyGeneration).filter(
        TurbineHourlyGeneration.daily_report_id.in_(report_ids)
    ).order_by(asc(TurbineHourlyGeneration.daily_report_id)).all()
    
    # Create Excel file
    excel_file = create_excel_report(daily_reports, turbine_stats, hourly_generations)
    
    # Generate filename
    filename = f"plant_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    ) 